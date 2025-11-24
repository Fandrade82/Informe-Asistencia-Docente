from flask import Flask, request, send_file, render_template
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')  # Usa el HTML con Bootstrap en /templates

@app.route('/procesar', methods=['POST'])
def procesar():
    try:
        # Leer archivo subido
        file = request.files['file']
        df = pd.read_excel(file, engine="openpyxl")

        # Validar columnas requeridas
        required_cols = [
            "Fecha", "Semana", "Hora1", "Hora2", "Hora3", "Hora4",
            "Tiempo total de trabajo", "Departamento", "Apellido y Nombre"
        ]
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            return f"Error: Faltan columnas en el archivo: {', '.join(missing)}", 400

        # Eliminar sábados y domingos
        df = df[~df["Semana"].isin(["sábado", "domingo"])]

        # Crear libro de salida
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Asistencia"

        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        # Agrupar por docente
        grouped = df.groupby("Apellido y Nombre")

        row = 1
        for docente, data in grouped:
            # Título con nombre del docente
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=docente)
            cell.font = Font(bold=True, size=14)
            cell.alignment = center_align
            row += 1

            # Encabezado
            headers = ["Fecha", "Departamento", "Hora1", "Hora2", "Hora3", "Hora4", "Tiempo total", "Observación"]
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            row += 1

            # Reglas de validación
            dept = data["Departamento"].iloc[0]
            for _, r in data.iterrows():
                fecha_dia = f"{r['Fecha']} - {r['Semana']}"
                hora1, hora2, hora3, hora4 = r["Hora1"], r["Hora2"], r["Hora3"], r["Hora4"]
                tiempo_total = r["Tiempo total de trabajo"]
                observacion = ""

                # Validar marcaciones
                if dept in ["ADMIN MATUTINA", "ADMIN VESPERTINA"]:
                    if pd.isna(hora3) or pd.isna(hora4):
                        observacion = "No marca"
                elif dept in ["DOC. MATUTINA", "DOC. VESPERTINA", "DOC. NOCTURNA"] and r["Semana"] == "jueves":
                    if pd.isna(hora3) or pd.isna(hora4):
                        observacion = "No marca"

                # Escribir fila
                values = [fecha_dia, dept, hora1, hora2, hora3, hora4, tiempo_total, observacion]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.alignment = center_align
                    if observacion:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                row += 1

            row += 2  # Espacio entre docentes

        # Ajustar ancho de columnas
        for col_idx in range(1, 9):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True,
                         download_name="Reporte_Asistencia.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        return f"Error interno: {str(e)}", 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
